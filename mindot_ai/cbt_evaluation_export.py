"""감사 가능한 CBT 평가 JSONL을 Dashboard 포함 Excel로 내보냅니다."""

from __future__ import annotations

import argparse
import json
from collections import Counter, defaultdict
from collections.abc import Iterable, Mapping
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SHEET_COLUMNS: dict[str, tuple[str, ...]] = {
    "Summary": (
        "metric", "anonymous_version", "actual_version", "scope", "count",
        "mean", "sample_stddev", "min", "max", "pass_status", "notes",
    ),
    "Case Results": (
        "case_id", "anonymous_version", "actual_version", "situation_family",
        "vulnerability_type", "expected_decision", "actual_decision",
        "question_purpose", "semantic_route_type", "question_goal",
        "preface_required", "preface_goal", "example_options",
        "grounding_question_codes", "avoid_topics", "writer_input",
        "writer_output", "final_response", "overall_score", "safety_result",
        "fallback_used", "validation_failure", "corrective_retry_count",
        "latency_ms", "agent_input_tokens", "agent_output_tokens",
        "writer_input_tokens", "writer_output_tokens", "total_tokens",
        "model_call_count", "critical_failure_count", "grader_status",
        "grader_note",
    ),
    "Rubric Scores": (
        "anonymous_version", "actual_version", "case_id", "situation_family",
        "vulnerability_type", "rubric_area", "score", "max_score",
        "reason", "critical_failure", "grader_note",
    ),
    "Confirmation Details": (
        "anonymous_version", "actual_version", "case_id",
        "distortion_candidates", "evidence_for_code", "evidence_for_excerpt",
        "evidence_against_code", "evidence_against_excerpt",
        "alternative_view_code", "alternative_view_excerpt",
        "acknowledgement_code", "acknowledgement_excerpt", "balanced_thought",
        "final_display_message", "deterministic_four_domain_ready",
    ),
    "Safety": (
        "anonymous_version", "actual_version", "case_id", "situation_family",
        "safety_variant", "expected_safety_action", "actual_safety_action",
        "expected_reason", "actual_reason", "expected_evidence", "actual_evidence",
        "current_user_match", "currentness_match", "false_positive",
        "false_negative", "passed", "gate_status", "grader_note",
    ),
    "Failures": (
        "anonymous_version", "actual_version", "case_id", "situation_family",
        "vulnerability_type", "failure_type", "severity", "detected",
        "evidence_excerpt", "linked_rubric_area", "grader_note",
    ),
    "Failure Audit": (
        "case_id", "anonymous_version", "actual_version", "failure_candidate",
        "source", "evidence_excerpt", "linked_rubric_area", "raw_candidate",
        "consistency_valid", "audited_failure", "rejection_reason",
    ),
    "Token Usage": (
        "anonymous_version", "actual_version", "case_id", "path_type",
        "operation", "agent_input_tokens", "agent_output_tokens",
        "writer_input_tokens", "writer_output_tokens", "total_tokens",
        "model_call_count", "corrective_retry_count", "fallback_used",
        "confirmation_tool_exposed",
    ),
    "Latency": (
        "anonymous_version", "actual_version", "case_id", "path_type",
        "operation", "latency_ms", "model_call_count", "fallback_used",
        "validation_failure",
    ),
    "Evaluation Context": (
        "case_id", "anonymous_version", "actual_version", "situation_family",
        "vulnerability_type", "automatic_thought", "latest_interaction",
        "latest_user_intent_hint", "previous_questions", "question_answers",
        "blocked_routes", "blocked_route_families",
        "resolved_but_irrelevant_topics", "expected_decision",
        "allowed_question_meanings", "forbidden_question_meanings",
        "semantic_route_definitions", "safety_candidates",
        "confirmation_allowed", "four_domain_coverage",
    ),
    "Metadata": ("key", "value"),
}

OUTPUT_SHEETS = ("Dashboard", *SHEET_COLUMNS)
Q2_COLOR = "4472C4"
Q4R_COLOR = "ED7D31"
PASS_COLOR = "70AD47"
FAIL_COLOR = "C00000"
NAVY = "1F4E78"
LIGHT_BG = "F4F6F8"


def append_jsonl(path: Path, sheet: str, row: Mapping[str, Any]) -> None:
    if sheet not in SHEET_COLUMNS:
        raise ValueError(f"Unknown evaluation sheet: {sheet}")
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as stream:
        stream.write(
            json.dumps({"sheet": sheet, "row": dict(row)}, ensure_ascii=False)
            + "\n"
        )
        stream.flush()


def read_jsonl(path: Path) -> dict[str, list[dict[str, Any]]]:
    rows = {sheet: [] for sheet in SHEET_COLUMNS}
    with path.open("r", encoding="utf-8") as stream:
        for line_number, line in enumerate(stream, start=1):
            if not line.strip():
                continue
            payload = json.loads(line)
            sheet = payload.get("sheet")
            row = payload.get("row")
            if sheet not in SHEET_COLUMNS or not isinstance(row, dict):
                raise ValueError(f"Invalid JSONL record at line {line_number}")
            rows[sheet].append(row)
    return rows


def _cell_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def _style_sheet(sheet: Any, column_count: int) -> None:
    header_fill = PatternFill("solid", fgColor=NAVY)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 30
    for column in range(1, column_count + 1):
        letter = get_column_letter(column)
        values = [
            sheet.cell(row=row, column=column).value
            for row in range(1, min(sheet.max_row, 200) + 1)
        ]
        longest = max(
            (len(str(value)) for value in values if value is not None),
            default=10,
        )
        sheet.column_dimensions[letter].width = min(max(longest + 2, 12), 48)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _summary_value(rows: list[dict[str, Any]], metric: str, version: str) -> float:
    for row in rows:
        if (
            str(row.get("metric", "")).upper() == metric.upper()
            and str(row.get("actual_version", "")).upper() == version.upper()
        ):
            return float(row.get("mean") or 0)
    return 0.0


def _version_order(rows: list[dict[str, Any]]) -> tuple[str, str]:
    versions = {
        str(row.get("actual_version"))
        for row in rows
        if row.get("actual_version")
    }
    q2 = next((item for item in versions if item.upper() == "Q2"), "Q2")
    q4r = next((item for item in versions if item.upper() == "Q4R"), "Q4R")
    return q2, q4r


def _metadata(rows: list[dict[str, Any]]) -> dict[str, Any]:
    return {str(row.get("key")): row.get("value") for row in rows}


def _set_chart_colors(chart: BarChart) -> None:
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = Q2_COLOR
    if len(chart.series) > 1:
        chart.series[1].graphicalProperties.solidFill = Q4R_COLOR


def _bar_chart(
    sheet: Any,
    *,
    title: str,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
    anchor: str,
    y_title: str,
) -> None:
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = "Category"
    chart.height = 8.2
    chart.width = 14.5
    data = Reference(
        sheet,
        min_col=min_col,
        max_col=max_col,
        min_row=min_row,
        max_row=max_row,
    )
    categories = Reference(
        sheet,
        min_col=min_col - 1,
        min_row=min_row + 1,
        max_row=max_row,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.legend.position = "b"
    _set_chart_colors(chart)
    sheet.add_chart(chart, anchor)


def _dashboard_source(rows: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    summary = rows["Summary"]
    cases = rows["Case Results"]
    audits = rows["Failure Audit"]
    tokens = rows["Token Usage"]
    latency_rows = rows["Latency"]
    q2, q4r = _version_order(summary)
    quality_metrics = (
        "OVERALL CBT QUALITY",
        "QUESTION TURN QUALITY",
        "CONFIRMATION QUALITY",
        "SAFETY ACCURACY",
    )
    quality = [
        (metric, _summary_value(summary, metric, q2), _summary_value(summary, metric, q4r))
        for metric in quality_metrics
    ]
    vulnerabilities = (
        "first_question",
        "no_direct_evidence",
        "relevance_feedback",
        "repetition_feedback",
        "example_request",
        "negated_or_non_current_safety",
        "confirmation_ready",
    )
    vulnerability_values: list[tuple[str, float, float]] = []
    for vulnerability in vulnerabilities:
        values: dict[str, list[float]] = defaultdict(list)
        for row in cases:
            if (
                row.get("vulnerability_type") == vulnerability
                and row.get("overall_score") is not None
                and row.get("grader_status", "VALID") != "GRADER_INVALID"
            ):
                values[str(row.get("actual_version"))].append(float(row["overall_score"]))
        vulnerability_values.append(
            (
                vulnerability,
                sum(values[q2]) / len(values[q2]) if values[q2] else 0,
                sum(values[q4r]) / len(values[q4r]) if values[q4r] else 0,
            )
        )
    paired: dict[str, dict[str, float]] = defaultdict(dict)
    for row in cases:
        if row.get("overall_score") is not None:
            paired[str(row.get("case_id"))][str(row.get("actual_version"))] = float(row["overall_score"])
    changes = Counter({"improved": 0, "tied": 0, "declined": 0})
    for values in paired.values():
        if q2 not in values or q4r not in values:
            continue
        difference = values[q4r] - values[q2]
        changes["improved" if difference > 0 else "declined" if difference < 0 else "tied"] += 1
    failure_groups = {
        "semantic repeat": {"semantic_repeat"},
        "direct-evidence re-ask": {"direct_evidence_reask"},
        "hidden third-party state": {"hidden_third_party_state"},
        "route mismatch": {"route_mismatch"},
        "confirmation failures": {
            "confirmation_without_four_domains", "confirmation_evidence_omitted",
            "confirmation_too_abstract", "confirmation_distortion_not_specific",
        },
        "safety failures": {"safety_failure"},
    }
    failures: list[tuple[str, int, int]] = []
    for label, names in failure_groups.items():
        counts = Counter()
        for row in audits:
            if row.get("audited_failure") in {True, "true", "TRUE", 1} and row.get("failure_candidate") in names:
                counts[str(row.get("actual_version"))] += 1
        failures.append((label, counts[q2], counts[q4r]))
    efficiency: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    case_count = Counter(str(row.get("actual_version")) for row in cases)
    for row in tokens:
        version = str(row.get("actual_version"))
        efficiency[version]["total_tokens"] += float(row.get("total_tokens") or 0)
        efficiency[version]["calls"] += float(row.get("model_call_count") or 0)
    latency_count: Counter[str] = Counter()
    for row in latency_rows:
        version = str(row.get("actual_version"))
        if row.get("latency_ms") is not None:
            efficiency[version]["latency"] += float(row["latency_ms"])
            latency_count[version] += 1
    for version in (q2, q4r):
        efficiency[version]["average_tokens"] = (
            efficiency[version]["total_tokens"] / case_count[version] if case_count[version] else 0
        )
        efficiency[version]["average_latency_ms"] = (
            efficiency[version]["latency"] / latency_count[version] if latency_count[version] else 0
        )
    return {
        "q2": q2, "q4r": q4r, "quality": quality,
        "vulnerabilities": vulnerability_values, "changes": changes,
        "failures": failures, "efficiency": efficiency,
    }


def _build_dashboard(workbook: Workbook, rows: dict[str, list[dict[str, Any]]]) -> None:
    sheet = workbook.create_sheet("Dashboard", 0)
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.freeze_panes = "A11"
    for column in range(1, 13):
        sheet.column_dimensions[get_column_letter(column)].width = 14
    for column in range(14, 20):
        sheet.column_dimensions[get_column_letter(column)].hidden = True
    metadata = _metadata(rows["Metadata"])
    source = _dashboard_source(rows)
    q2, q4r = source["q2"], source["q4r"]
    sheet.merge_cells("A1:L2")
    sheet["A1"] = "CBT Agent Q2 vs Q4R — Audited Evaluation"
    sheet["A1"].font = Font(size=22, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.merge_cells("A3:L3")
    sheet["A3"] = (
        f"UTC {metadata.get('evaluation_utc', '')}  |  "
        f"Set {metadata.get('evaluation_set', 'CBT_AGENT_STRESS_V2_AUDITED')}  |  "
        f"Q2 {metadata.get('q2_commit', '')}  |  Q4R {metadata.get('q4r_commit', '')}"
    )
    sheet["A3"].alignment = Alignment(horizontal="center")
    sheet["A3"].font = Font(color="44546A", italic=True)
    quality_lookup = {item[0]: item for item in source["quality"]}
    audited_failures = {
        version: sum(
            1 for row in rows["Failure Audit"]
            if str(row.get("actual_version")) == version
            and row.get("audited_failure") in {True, "true", "TRUE", 1}
        )
        for version in (q2, q4r)
    }
    paired_diff = quality_lookup["OVERALL CBT QUALITY"][2] - quality_lookup["OVERALL CBT QUALITY"][1]
    cards = (
        ("Q2 OVERALL", quality_lookup["OVERALL CBT QUALITY"][1], Q2_COLOR),
        ("Q4R OVERALL", quality_lookup["OVERALL CBT QUALITY"][2], Q4R_COLOR),
        ("PAIRED Δ", paired_diff, PASS_COLOR if paired_diff >= 0 else FAIL_COLOR),
        ("QUESTION Q4R", quality_lookup["QUESTION TURN QUALITY"][2], Q4R_COLOR),
        ("CONFIRM Q4R", quality_lookup["CONFIRMATION QUALITY"][2], Q4R_COLOR),
        ("SAFETY Q4R", quality_lookup["SAFETY ACCURACY"][2], PASS_COLOR),
        ("AUDITED FAIL Q4R", audited_failures[q4r], FAIL_COLOR),
        ("TOTAL TOKENS Q4R", source["efficiency"][q4r]["total_tokens"], Q4R_COLOR),
        ("AVG LATENCY Q4R", source["efficiency"][q4r]["average_latency_ms"], Q4R_COLOR),
    )
    thin = Side(style="thin", color="D9E1F2")
    for index, (label, value, color) in enumerate(cards):
        column = 1 + (index % 5) * 2
        row = 5 + (index // 5) * 3
        sheet.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column + 1)
        sheet.merge_cells(start_row=row + 1, start_column=column, end_row=row + 1, end_column=column + 1)
        label_cell = sheet.cell(row=row, column=column, value=label)
        value_cell = sheet.cell(row=row + 1, column=column, value=value)
        for cell in (label_cell, value_cell):
            cell.fill = PatternFill("solid", fgColor=LIGHT_BG)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        label_cell.font = Font(size=9, bold=True, color="44546A")
        value_cell.font = Font(size=16, bold=True, color=color)
        value_cell.number_format = "0.00"
    start_col = 14
    row = 1
    sheet.cell(row=row, column=start_col, value="Quality")
    sheet.cell(row=row, column=start_col + 1, value=q2)
    sheet.cell(row=row, column=start_col + 2, value=q4r)
    for label, q2_value, q4r_value in source["quality"]:
        row += 1
        sheet.cell(row=row, column=start_col, value=label)
        sheet.cell(row=row, column=start_col + 1, value=q2_value)
        sheet.cell(row=row, column=start_col + 2, value=q4r_value)
    quality_end = row
    row = 8
    sheet.cell(row=row, column=start_col, value="Vulnerability")
    sheet.cell(row=row, column=start_col + 1, value=q2)
    sheet.cell(row=row, column=start_col + 2, value=q4r)
    for label, q2_value, q4r_value in source["vulnerabilities"]:
        row += 1
        sheet.cell(row=row, column=start_col, value=label)
        sheet.cell(row=row, column=start_col + 1, value=q2_value)
        sheet.cell(row=row, column=start_col + 2, value=q4r_value)
    vulnerability_end = row
    row = 18
    sheet.cell(row=row, column=start_col, value="Paired result")
    sheet.cell(row=row, column=start_col + 1, value="Count")
    for label in ("improved", "tied", "declined"):
        row += 1
        sheet.cell(row=row, column=start_col, value=label)
        sheet.cell(row=row, column=start_col + 1, value=source["changes"][label])
    changes_end = row
    row = 24
    sheet.cell(row=row, column=start_col, value="Audited failure")
    sheet.cell(row=row, column=start_col + 1, value=q2)
    sheet.cell(row=row, column=start_col + 2, value=q4r)
    for label, q2_value, q4r_value in source["failures"]:
        row += 1
        sheet.cell(row=row, column=start_col, value=label)
        sheet.cell(row=row, column=start_col + 1, value=q2_value)
        sheet.cell(row=row, column=start_col + 2, value=q4r_value)
    failures_end = row
    row = 33
    sheet.cell(row=row, column=start_col, value="Token metric")
    sheet.cell(row=row, column=start_col + 1, value=q2)
    sheet.cell(row=row, column=start_col + 2, value=q4r)
    for label, key in (("Total tokens", "total_tokens"), ("Average tokens", "average_tokens")):
        row += 1
        sheet.cell(row=row, column=start_col, value=label)
        sheet.cell(row=row, column=start_col + 1, value=source["efficiency"][q2][key])
        sheet.cell(row=row, column=start_col + 2, value=source["efficiency"][q4r][key])
    token_end = row
    row = 38
    sheet.cell(row=row, column=start_col, value="Latency")
    sheet.cell(row=row, column=start_col + 1, value=q2)
    sheet.cell(row=row, column=start_col + 2, value=q4r)
    sheet.cell(row=row + 1, column=start_col, value="Average latency ms")
    sheet.cell(row=row + 1, column=start_col + 1, value=source["efficiency"][q2]["average_latency_ms"])
    sheet.cell(row=row + 1, column=start_col + 2, value=source["efficiency"][q4r]["average_latency_ms"])
    row = 42
    sheet.cell(row=row, column=start_col, value="Calls")
    sheet.cell(row=row, column=start_col + 1, value=q2)
    sheet.cell(row=row, column=start_col + 2, value=q4r)
    sheet.cell(row=row + 1, column=start_col, value="Model calls")
    sheet.cell(row=row + 1, column=start_col + 1, value=source["efficiency"][q2]["calls"])
    sheet.cell(row=row + 1, column=start_col + 2, value=source["efficiency"][q4r]["calls"])
    _bar_chart(sheet, title="Quality comparison", min_row=1, max_row=quality_end, min_col=15, max_col=16, anchor="A12", y_title="Score")
    _bar_chart(sheet, title="Vulnerability averages", min_row=8, max_row=vulnerability_end, min_col=15, max_col=16, anchor="G12", y_title="Score")
    _bar_chart(sheet, title="Paired outcomes", min_row=18, max_row=changes_end, min_col=15, max_col=15, anchor="A29", y_title="Cases")
    _bar_chart(sheet, title="Audited failures", min_row=24, max_row=failures_end, min_col=15, max_col=16, anchor="G29", y_title="Count")
    _bar_chart(sheet, title="Token efficiency", min_row=33, max_row=token_end, min_col=15, max_col=16, anchor="A46", y_title="Tokens")
    _bar_chart(sheet, title="Average latency", min_row=38, max_row=39, min_col=15, max_col=16, anchor="G46", y_title="Milliseconds")
    _bar_chart(sheet, title="Model calls", min_row=42, max_row=43, min_col=15, max_col=16, anchor="A63", y_title="Calls")
    sheet.print_area = "A1:L80"


def export_workbook(jsonl_path: Path, output_path: Path) -> Path:
    rows_by_sheet = read_jsonl(jsonl_path)
    workbook = Workbook()
    workbook.remove(workbook.active)
    _build_dashboard(workbook, rows_by_sheet)
    for sheet_name, columns in SHEET_COLUMNS.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(list(columns))
        for row in rows_by_sheet[sheet_name]:
            sheet.append([_cell_value(row.get(column)) for column in columns])
        _style_sheet(sheet, len(columns))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def validate_workbook(
    path: Path,
    *,
    expected_row_counts: Mapping[str, int] | None = None,
) -> dict[str, int]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    if tuple(workbook.sheetnames) != OUTPUT_SHEETS:
        raise ValueError(f"Unexpected sheets: {workbook.sheetnames}")
    dashboard = workbook["Dashboard"]
    if len(dashboard._charts) < 5:
        raise ValueError("Dashboard must contain at least five charts")
    for chart in dashboard._charts:
        if not chart.series:
            raise ValueError(f"Dashboard chart has no series: {chart.title}")
        for series in chart.series:
            formula = getattr(getattr(series, "val", None), "numRef", None)
            if formula is None or not formula.f:
                raise ValueError(f"Dashboard chart has no source range: {chart.title}")
    row_counts: dict[str, int] = {"Dashboard": 0}
    for sheet_name, columns in SHEET_COLUMNS.items():
        sheet = workbook[sheet_name]
        headers = tuple(cell.value for cell in sheet[1])
        if headers != columns:
            raise ValueError(f"Unexpected columns in {sheet_name}: {headers}")
        row_counts[sheet_name] = max(sheet.max_row - 1, 0)
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("#"):
                    raise ValueError(f"Formula error marker in {sheet_name}!{cell.coordinate}")
    workbook.close()
    if expected_row_counts is not None:
        for sheet_name, expected in expected_row_counts.items():
            if row_counts.get(sheet_name) != expected:
                raise ValueError(
                    f"Unexpected row count for {sheet_name}: "
                    f"{row_counts.get(sheet_name)} != {expected}"
                )
    return row_counts


def write_rows(path: Path, sheet: str, rows: Iterable[Mapping[str, Any]]) -> None:
    for row in rows:
        append_jsonl(path, sheet, row)


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("jsonl", type=Path)
    parser.add_argument("output", type=Path)
    return parser.parse_args()


if __name__ == "__main__":
    args = _parse_args()
    generated = export_workbook(args.jsonl, args.output)
    counts = validate_workbook(generated)
    print(json.dumps({"path": str(generated), "rows": counts}, ensure_ascii=False))
